from openpyxl import load_workbook
from openpyxl.styles import PatternFill


#returns cell values, this two different cells in a dictionary to get the values
def returnCellVal(cell, desc_cell, worksheet):
    
    cellValues = {'sku': worksheet[cell].value, 'descrip': worksheet[desc_cell].value}
    return cellValues

#the function below does not work
def changeColorCell(cell, worksheet):
    for rows in worksheet.iter_rows(min_row = cell, max_row = 1, min_col=1):
        for cell in rows:
            cell.fill = PatternFill(bgColor= "0000FF00", fill_type = "solid")

if __name__ == '__main__':

    workbook = load_workbook(input('Enter name of excel file:\n')+ '.xlsx')
    worksheet = workbook[input('Enter name of worksheet:\n')]

    values = returnCellVal('A2', 'D2', worksheet)
    print(values['sku'] + ' ' + values['descrip'])
    changeColorCell(1, worksheet)
