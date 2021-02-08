import openpyxl
import re
from openpyxl import load_workbook


def GetWorkBooks():
    #creates new workbook to save rows in
    newWorkbook1 = openpyxl.Workbook()
    newWorkbook = openpyxl.Workbook()
    ws1 = newWorkbook1.active 
    ws = newWorkbook.active

    #gets the first workbook that will be analyzed
    workbook1 = openpyxl.load_workbook( 'ExcelFiles/' + input('Enter name of first excel file:\n')+ '.xlsx')
    worksheet1 = workbook1[input('Enter name of first worksheet:\n')]

    #gets second workbook to be analyzed against the first
    workbook2 = openpyxl.load_workbook('ExcelFiles/' + input('Enter name of second excel file:\n') + '.xlsx')
    worksheet2 = workbook2[input('Enter name of second worksheet\n')]
    rowNum = 1

    #starts a loop to grab the first sku
    for row1 in range(2, worksheet1.max_row):
            sku1 = worksheet1['A' + str(row1)].value
            sku1 = re.sub(".*\\s", "", sku1)
            print(sku1)

            #second loop to check against the sku from the first loop
            for row2 in range(2, worksheet2.max_row+1):
                sku2 = worksheet2['C' + str(row2)].value
                if(sku1 == sku2):
                    print(sku1 + " " + sku2 + " are the same. Starting the data entry process.")

                    for nCol1 in range(1, worksheet1.max_column +1 ):
                        rowData = worksheet1.cell(row = row1, column = nCol1)
                        ws1.cell(row = rowNum, column = nCol1).value = rowData.value
                    
                    #loop to get all the values from the columns
                    for nCol in range(1, worksheet2.max_column + 1): 
                        rowData= worksheet2.cell(row= row2, column= nCol)
                        ws.cell(row = rowNum, column= nCol).value = rowData.value
                    print('Row has been entered')
                    rowNum = rowNum + 1

    #saves the workbook once we drop out of all the loops                    
    newWorkbook.save('temp.xlsx')
    newWorkbook1.save('temp1.xlsx')        

if __name__ == '__main__':

    GetWorkBooks()
