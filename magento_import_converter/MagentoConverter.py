import openpyxl
from openpyxl import load_workbook
import re

def OpenWorkbook():
    #opens the workbook
    toConvertWb = openpyxl.load_workbook('Backup/ExcelSheets/' + input('enter the name of workbook\n') + '.xlsx')
    toConvertWs = toConvertWb[input('Enter name of the worksheet to open:\n')]

    useOfScript = input('Would you like to (1)convert for import or (2) replace skus? type 1 or 2 respectively\n')

    #if user wants to convert for import
    if int(useOfScript) == 1:
        #opens the base file used for import in Magento
        importWb = openpyxl.load_workbook('Backup/ExcelSheets/catalog_product.xlsx')
        importWs = importWb['catalog_product']
        rows = 2
        #loops throughout the entire workbook to get the correct information
        for row in range(2, toConvertWs.max_row + 1):
            #checks that it is base unit before pulling information
            if(toConvertWs['G' + str(row)].value != None):
                if (toConvertWs['G' + str(row)].value.lower() == 'base_unit_or_each'):
                    #getting all information from row
                    sku = toConvertWs['C' + str(row)].value
                    itemName = toConvertWs['O' + str(row)].value
                    description = ""
                    productOnline = '1'
                    additionalInfo = ''
                    weight = toConvertWs['AE' + str(row)].value
                    picture = ''

                    #Checks for a description
                    if(toConvertWs['P' + str(row)].value != None):
                        description = '<p>' + toConvertWs['P' + str(row)].value + '</p>'
                        
                    if(weight == None):
                        weight = ''
                        productOnline= '2'
                    
                    #checks to see if item name is none, if so changes online status to not online
                    if(itemName == None):
                        itemName = ''
                        productOnline = '2'

                    #checks for the brand of the product
                    if(sku[0:3] == 'ASM'):
                        additionalInfo = 'brand=Lenox'

                    if(sku[0:3] == 'B&D'):
                        additionalInfo = 'brand=Dewalt'

                    if(sku[0:3] == 'BOS'):
                        additionalInfo = 'brand=Bostitch'

                    if(sku[0:3] == 'ELC'):
                        additionalInfo = 'brand=Elco'

                    if(sku[0:3] == 'IRW'):
                        additionalInfo = 'brand=Irwin'

                    if(sku[0:3]== 'PCT' or sku[0:3] == 'RAW'):
                        additionalInfo = 'brand=Powers'

                    if(sku[0:3]== 'SBD' or sku[0:3] == 'STA'):
                        additionalInfo = 'brand=Stanley'

                    #checks for a picture on the row
                    if(toConvertWs['AG' + str(row)].value != None):
                        picture = toConvertWs['AG' + str(row)].value
                    elif(toConvertWs['AH' + str(row)].value != None):
                        picture = toConvertWs['AH' + str(row)].value
                    else:
                        #remove string for no picture
                        picture = ''
                        productOnline = '2'
                
                    importWs = ConvertToImport(sku, productOnline, itemName, description, weight, picture, additionalInfo, rows, importWs)
                    rows = rows + 1
            importWb.save('text.xlsx')

    #if user wants to replace sku's
    elif int(useOfScript) == 2:
        wbToUse = openpyxl.load_workbook('Backup/ExcelSheets/' + input('enter the name of workbook:\n') + '.xlsx')
        wsToUse =  wbToUse[input('Enter name of the worksheet to open:\n')]
        wsToUse = ReplaceSku(toConvertWs, wsToUse)
        wbToUse.save('ConvertedFile.xlsx')
    else:
        print("Program will now exit.")
                                     
def ReplaceSku(wsUse, toConvWs):
     
    #replaces the Sku's between two sheets
    print("to replace Sku's")
    for row in range(2, wsUse.max_row + 1):
            sku = wsUse['A' + str(row)].value
            if(sku == None):
                break
            print(sku)
            trimSku = re.sub(".*\\s", "", sku)
            
            for row2 in range(2, toConvWs.max_row + 1):
                sku2 = toConvWs['C' + str(row2)].value
                if(trimSku == sku2):
                    print('Replacing Sku')
                    toConvWs['C' + str(row2)].value = sku
    return toConvWs

def ConvertToImport(sku, productStatus, itemName, description, weight, picture, additionalInfo, rows, worksheet):
    #TODO: Add function to convert to import sheet
    
    print("Sku: " + sku + " Item Name: " + itemName + " Product Status: " + productStatus + " Weight: " + weight + " Picture: " + picture)
    category = 'Default/Products'
    attSetCol = 'Default'
    productType = 'simple'
    productWebsite = 'base'
    taxableGood = 'Taxable Goods'
    visibility = 'Catalog, Search'
    worksheet['A' + str(rows)].value = sku
    worksheet['B' + str(rows)].value = productWebsite
    worksheet['C' + str(rows)].value = attSetCol
    worksheet['D' + str(rows)].value = productType
    worksheet['E' + str(rows)].value = category
    worksheet['G' + str(rows)].value = itemName
    worksheet['I' + str(rows)].value = description
    worksheet['K' + str(rows)].value = productStatus
    worksheet['J' + str(rows)].value = weight
    worksheet['L' + str(rows)].value = taxableGood
    worksheet['M' + str(rows)].value = visibility
    worksheet['T' + str(rows)].value = additionalInfo
    for nCol in range(15, 20):
        worksheet.cell(row = rows, column = nCol).value = picture
    return worksheet
    
if __name__ == "__main__":

    OpenWorkbook()
