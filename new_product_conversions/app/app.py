import openpyxl
import re
import sys
from openpyxl import load_workbook, Workbook
from Classes.ExcelSheetObject import ExcelSheet

if __name__ == '__main__':

    #creates a new workbook
    wb = openpyxl.Workbook()
    
    while True:
        try:
            sheet = ExcelSheet(input('Please enter excel sheet name:\n'), input('Enter Sku Identification to add to beginning of SKU:\n'), input('Enter starting row:\n'))
            workbook = openpyxl.load_workbook('../ExcelFiles/' + sheet.sheet_name)
            break
        except:
            answer = input("This workbook was not found! try again, Enter Q to quit the program, otherwise hit enter to continue.\n")
            if(answer == 'Q'):
                sys.exit('You have quit the program!')

    #initializing dictionary for each column
    colDiction= {}
    
    #following loop will return a dictionary of column letters
    numOfCols = sheet.totalCols()

    #add column to discitonary per user input
    for i in range(1,numOfCols):
        colDiction['col{0}'.format(i)] = sheet.returnCols()

    #iterates through each sheet in the worbook
    for sheets in workbook.worksheets:
        
        #creates new worksheet in the copy workbook
        ws = wb.create_sheet(sheets.title)
        
        #iterates through the columns
        for i in range(1,numOfCols):
            #row for new worksheet
            rowForWs = 1
            
            #iterates through the rows
            for row in range(sheet.starting_row, sheets.max_row + 1):
                ws.cell(row = rowForWs, column = i).value = sheets[colDiction['col{0}'.format(i)] + str(row)].value
                rowForWs += 1
                
    #iterates through the new worksheet and uses the identifier to change the sku in the workbook
    for ws in wb.worksheets:
        for i in range(1, ws.max_row + 1):
            fixedSku = sheet.manu_ident + " " + str(ws.cell(row = i, column = 1).value)
            ws.cell(row=i, column = 1).value = fixedSku

    wb.save('testSheet.xlsx')       
    print(sheet.sheet_name)
