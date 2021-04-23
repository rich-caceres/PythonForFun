import openpyxl
import re
import sys
from openpyxl import load_workbook, Workbook
from Classes.ExcelSheetObject import ExcelSheet


def workbookCheck():

    columnLetter = None
    while True:
        try:
            magentoSheet = ExcelSheet(input('Please enter the name of the Magento exported excel sheet:\n'), '', input('Enter starting row:\n'))
            excelSheetToCheck = ExcelSheet(input('Please enter the name of the excel sheet to check:\n'), input('Enter initial identifier for manufacturer:\n'), input('Enter starting row:\n'))
            mageXcelWorkbook = openpyxl.load_workbook('../ExcelFiles/' + magentoSheet.sheet_name)
            mageXcelSheet = mageXcelWorkbook.active
            XcelWorkbook = openpyxl.load_workbook('../ExcelFiles/' + excelSheetToCheck.sheet_name)
            XcelWorksheet = XcelWorkbook.active
            break
        except:
            answer = input("This work book was not found! press enter to try again. Enter the letter Q to quit the program")
            if(answer == 'Q'):
                sys.exit('You have quit the program!')

    productsNotInCatalog= {}
    productNotInCatalogKey = 1
    for row1 in range(excelSheetToCheck.starting_row, XcelWorksheet.max_row + 1):
        if columnLetter is None:
            columnLetter= input('Please enter the column letter containing SKU on the sheet the contains the cross sell items:\n')
        Sku1 = XcelWorksheet['B' + str(row1)].value
        upsell_products = {}
        for row2 in range(magentoSheet.starting_row, mageXcelSheet.max_row + 1):
            Sku2 = str(mageXcelSheet['A' + str(row2)].value)
            Sku2 = re.sub(".*\\s", "", Sku2)
            if(Sku1 == Sku2):
                print(f'Entered if Statement {Sku1} == {Sku2}')#testing
                
                for cols in range(5, XcelWorksheet.max_column+1):
                    
                    if XcelWorksheet.cell(row = row1, column = cols).value is not None:
                        upsell_products['{0}'.format(cols)] =excelSheetToCheck.manu_ident + " " + XcelWorksheet.cell(row=row1, column=cols).value
                valueString = None
                position = None
                if bool(upsell_products):
                    #print('in second if for bool products')#testing
                    #print(upsell_products)#testing
                    for key in list(upsell_products):
                        #print(upsell_products)#testing
                        if productsNotInCatalog is not None:
                            for keys in list(productsNotInCatalog):
                                try:
                                    if(productsNotInCatalog[keys] == upsell_products[key]):
                                        upsell_products.pop(key)
                                        continue
                                except KeyError:
                                    print(f'The Key: {key} does not exist, returning to top of loop.')
                                    break
                        try:
                            if bool(upsell_products[key]):
                                print('passed test')
                        except KeyError:
                            print(f'The key: {key} does not exist, returning to the top of the loop.')
                            continue
                            
                        if not bool(upsell_products):
                            break
                        for rowCheck in range(magentoSheet.starting_row, mageXcelSheet.max_row+1):
                            #print(str(mageXcelSheet['A' + str(rowCheck)].value)[0:3]) #testing
                            if(str(mageXcelSheet['A' + str(rowCheck)].value)[0:3] != str(excelSheetToCheck.manu_ident)):
                                #print(upsell_products)#testing
                                
                                if(rowCheck == mageXcelSheet.max_row):
                                    print(f'Deleted {upsell_products[key]}')
                                    productsNotInCatalog[productNotInCatalogKey]= upsell_products[key]
                                    upsell_products.pop(key)
                                    productNotInCatalogKey += 1
                                    break
                                
                                continue
                            try:
                                if(upsell_products[key] == mageXcelSheet['A' + str(rowCheck)].value):
                                    print('values equal if')#testing
                                    break
                            except KeyError:
                                print(f'The key:{key} does not exist in the Dictionary.')
                                break
                                        
                    print(upsell_products)
                    for addKey in list(upsell_products):
                        if (valueString == None and position == None):
                            valueString = upsell_products[addKey]
                            position = '0'
                        else:
                            valueString = valueString+','+upsell_products[addKey]
                            position = position + ',' + '0'
                print(f'{valueString} \n {position}')
                if (valueString is not None and position is not None):
                    mageXcelSheet['BS' + str(row2)].value = valueString
                    mageXcelSheet['BT' + str(row2)].value = position
                #print(upsell_products)#testing
                break
                        
    mageXcelWorkbook.save('newMageWorkbook.xlsx')
    
if __name__ == '__main__':

    workbookCheck()
    print('Nothing is happening here yet')
