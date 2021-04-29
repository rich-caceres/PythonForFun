import openpyxl
import re
from openpyxl import load_workbook


def CheckMapp(numOfSheets):

    try:
        Workbook1 = openpyxl.load_workbook('ExcelFile/'+ input('Enter name of excel file from magento:\n') + '.xlsx', keep_vba=False)
        Worksheet1 = Workbook1.active

        Workbook2 = openpyxl.load_workbook('ExcelFile/' + input('Enter name of excel file with MAPP pricing:\n')+ '.xlsx', keep_vba = False)
        Worksheet2 = Workbook2[input('Enter the name of the worksheet with MAPP pricing:\n')]
    except:
        print("An Error occurred, try again")

    #Sets parameter to check for mapp and msrp
    startRowMapp = input('Enter the starting row for Mapp Sheet:\n')
    colWithMapp = input('Enter column letter with MAPP:\n')
    colWithMsrp = input('Enter column letter with MSRP:\n')
    startDate = input('Enter start date of MAPP:\n')
    endDate = input('Enter end date of MAPP:\n')

    #starts grabbing each sku from the first worksheet
    for row1 in range(2, Worksheet1.max_row):
        sku1 = Worksheet1['A' + str(row1)].value
        sku1 = re.sub(".*\\s", "", sku1)
        print(sku1)

        #grabs the next sku to check from the other worksheet
        for row2 in range(int(startRowMapp), Worksheet2.max_row):
            sku2 = Worksheet2['B' + str(row2)].value

            #checks both skus to start copying data over
            if(sku1 == sku2):
                print('Matching Sku found, adding mapp and msrp')
                mapp = Worksheet2[colWithMapp + str(row2)].value
                msrp = Worksheet2[colWithMsrp + str(row2)].value
                Worksheet1['E' + str(row1)].value = msrp
                Worksheet1['F' + str(row1)].value = mapp
                Worksheet1['G' + str(row1)].value = startDate
                Worksheet1['H' + str(row1)].value = endDate
                
    Workbook1.save(f'import{numOfSheets}.xlsx')

if __name__ == "__main__":

    answer = None
    numOfSheets = 1
    while True:

        try:
            answer = input("Would you like to convert a MAPP for import to Magento? y/n\n")
            if (answer.lower() == 'n'): break
            CheckMapp(numOfSheets)
            numOfSheet += 1
        except:
            print("An error ocurred, try again.")

