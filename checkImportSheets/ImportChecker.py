import openpyxl
import wget
import urllib
from urllib.error import HTTPError
from openpyxl import load_workbook

toImportWb = openpyxl.load_workbook('ExcelSheets/' + input('Enter name of the sheet that is staged to be imported\n')+ '.xlsx')
toImportWs = toImportWb.active

fromMagentoWb = openpyxl.load_workbook('ExcelSheets/' + input('Enter name of the sheet that was exported from Magento\n')+ '.xlsx')
fromMagentoWs = fromMagentoWb.active

url = ""
for row1 in range(2, toImportWs.max_row):
    sku1 = toImportWs['A' + str(row1)].value
    print(sku1)
    toImportWs['N' + str(row1)].value = '0'
    #wget downloads pictures from the source
    if (toImportWs['O' + str(row1)].value != None):
        try:
            if(toImportWs['O' + str(row1)].value != url):
                url = toImportWs['O' + str(row1)].value
                image_Filename = wget.download(url)
        except urllib.error.HTTPError as exception:
            print(exception)
            image_Filename = ""
            
        for nCol in range(15, 20):
            toImportWs.cell(row = row1, column = nCol).value = image_Filename
    
    for row2 in range(2, fromMagentoWs.max_row + 1):
        sku2 = fromMagentoWs['A' + str(row2)].value
        if(sku1 == sku2):
            #adds data to the cells that will be imported to Magento
            if(fromMagentoWs['C' + str(row2)].value != 'Default' or fromMagentoWs['C' + str(row2)].value != None):
                toImportWs['C' + str(row1)].value = fromMagentoWs['C' + str(row2)].value
                
                if(fromMagentoWs['E' + str(row2)].value != None):
                    toImportWs['E' + str(row1)].value = fromMagentoWs['E' + str(row2)].value

                if(fromMagentoWs['N' + str(row2)].value != None):
                    toImportWs['N' + str(row1)].value = fromMagentoWs['N' + str(row2)].value
                
                if(fromMagentoWs['G' + str(row2)].value != None):
                    toImportWs['G' + str(row1)].value = fromMagentoWs['G' + str(row2)].value

toImportWb.save('toImport.xlsx')
